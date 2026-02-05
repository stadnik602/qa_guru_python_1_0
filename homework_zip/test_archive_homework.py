import csv
import os
from io import TextIOWrapper, BytesIO
from pathlib import Path
from zipfile import ZipFile

from _pytest.fixtures import fixture
from openpyxl import load_workbook


def get_current_file_path():
    CURRENT_FILE = os.path.abspath(__file__)
    return CURRENT_FILE

def get_current_dir_path():
    current_dir = os.path.dirname(get_current_file_path())
    return current_dir

@fixture
def temp_dir():
    tmp_dir = os.path.join(get_current_dir_path(), "temp")
    return tmp_dir


if os.path.exists("temp"):
    print("the 'temp' folder exist")

# shutil.rmtree(os.path.join(CURRENT_DIR, "tmp2")) #remove directory

def test_create_zip_with_files(temp_dir):
    zip_path = Path(temp_dir) / "files.zip"

    if zip_path.exists():
        zip_path.unlink()

    pdf = Path(temp_dir) / "Python Testing With Pytest(Brian-Okken).pdf"
    xlsx = Path(temp_dir) / "APEX Check-list.xlsx"
    csv = Path(temp_dir) / "email.csv"

    with ZipFile(zip_path, "w") as zip_file:
        zip_file.write(pdf, arcname="Python Testing With Pytest(Brian-Okken).pdf")
        zip_file.write(xlsx, arcname="Checklist.xlsx")
        zip_file.write(csv, arcname="Email.csv")

    assert zip_path.exists()

    with ZipFile(zip_path) as zip_file:
        assert len(zip_file.namelist()) == 3
        assert set(zip_file.namelist()) == {
            "Python Testing With Pytest(Brian-Okken).pdf",
            "Checklist.xlsx",
            "Email.csv",
        }

def test_reading_files_from_zip(temp_dir):
    zip_path = Path(temp_dir) / "files.zip"
    with ZipFile(zip_path) as zip_file:
        names = zip_file.namelist()
        assert names == ["Python Testing With Pytest(Brian-Okken).pdf", "Checklist.xlsx", "Email.csv"]

        with zip_file.open("Python Testing With Pytest(Brian-Okken).pdf") as pdf_file:
            pdf_content = pdf_file.read()
            assert pdf_content.startswith(b"%PDF")

        with zip_file.open("Checklist.xlsx") as xlsx_file:
            xlsx_bytes = BytesIO(xlsx_file.read())
            wb = load_workbook(xlsx_bytes)
            sheet_names = wb.sheetnames
            assert "Billing" in sheet_names

        with zip_file.open("Email.csv") as csv_file:
            text = TextIOWrapper(csv_file, encoding="utf-8")
            reader = csv.DictReader(text)
            rows = list(reader)

            assert len(rows[1]) == 6
            assert any("Rog Swift Oled" in value for value in rows[1].values())
            assert int(rows[0]["LG"]) == 32
            assert rows[1]["Samsung"] == "Odyssey G8"
